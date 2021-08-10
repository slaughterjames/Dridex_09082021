#! /usr/bin/env python3
'''
 
'''

#python import
import sys
import os
import base64
import binascii
import struct
import re
from openpyxl import load_workbook

#programmer generated imports
from controller import controller
from fileio import fileio

'''
Usage()
Function: Display the usage parameters when called
'''
def Usage():
    print ('Usage: [required] --xls [optional] --debug --help')
    print ('Example: ./dridex_xls_09082021.py --xls sheet.xls --debug')
    print ('Required Arguments:')
    print ('--xls - sheet being examined')
    print ('Optional Arguments:')
    print ('--debug - Prints verbose logging to the screen to troubleshoot issues with a recon installation.')
    print ('--help - You\'re looking at it!')
    sys.exit(-1)
            
'''
Parse() - Parses program arguments
'''
def Parse(args):        
    option = ''

    print ('[*] Length Arguments: ' + str(len(args)))

    if (len(args) == 1):
        return -1

    print ('[*] Arguments: ')
    for i in range(len(args)):
        if args[i].startswith('--'):
            option = args[i][2:]
                
            if option == 'help':
                return -1

            if option == 'xls':
                CON.xls = args[i+1]
                print (option + ': ' + CON.xls)                              
                
            if option == 'debug':
                CON.debug = True
                print (option + ': ' + str(CON.debug))               
                                     
    #These are required params so length needs to be checked after all 
    #are read through         
    if (len(CON.xls) < 3):
        print ('[x] xls is a required argument.')
        print ('')
        return -1

    print ('')   
    
    return 0


'''
viaMail()
Function: - viaMail
'''
def viaMail(ws, max_rows):

    range_string = 'B169:V745'
    count = 0
    character = 0
    output = ''

    FLOG = fileio()

    print ('[*] In viaMail')

    col_start, col_end = re.findall("[A-Z]+", range_string)

    print ('[*] data_rows')
    data_rows = []
    for row in ws[range_string]:

        #print (cell.value.value for cell in row)
        data_rows.append([cell.value for cell in row])

    try:
        while count < max_rows-1:
            #print ('Count: ' + str(count))
            #print (data_rows[count])
            for val in data_rows[count]:
                if ((val != None)):# and (isinstance(val, int) == True)):
                    print ('val: ' + str(val))
                    character = round(val)# + 13000
                    print(isinstance(val, int))
                    print ('character: ' + str(character))
                    if ((character > 0) and (character < 130)):
                        try:
                            output += chr(character)
                        except ValueError as e:                   
                            print ('[x] Error - charcter: ' + str(e)) 
                character = 0
            FLOG.WriteLogFile('badfile.txt', output)
            output = ''
            count += 1
    except IndexError as e:                   
        print ('[x] Have reached the end of the index...')   

    return 0

'''
Execute()
Function: - Does the doing against a string
'''
def Execute():

    data_file = CON.xls

    # Load the entire workbook.
    CON.wb = load_workbook(data_file)

    # List all the sheets in the file.
    print("Found the following worksheets:")
    for sheetname in CON.wb.sheetnames:
        print('Sheetname: ' + sheetname)

    # Load one worksheet.
    ws = CON.wb['Sheet1']
    max_rows = ws.max_row
    all_rows = list(ws.rows)
    #sheet_ranges = CON.wb['range names']
    print ('Max Row: ' + str(max_rows))
    print(f"Found {len(all_rows)} rows of data.")

    viaMail(ws, max_rows)


    return 0

'''
Terminate()
Function: - Attempts to exit the program cleanly when called  
'''
     
def Terminate(exitcode):
    sys.exit(exitcode)

'''
This is the mainline section of the program and makes calls to the 
various other sections of the code
'''

if __name__ == '__main__':
    
    ret = 0

    #Stores our args
    CON = controller()
                   
    #Parses our args
    ret = Parse(sys.argv)

    #Something bad happened
    if (ret == -1):
        Usage()
        Terminate(ret)

    #Do the doing
    Execute()

    print ('')
    print ('[*] Program Complete')

    Terminate(0)
'''
END OF LINE
'''

